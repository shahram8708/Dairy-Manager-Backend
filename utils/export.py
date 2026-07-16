import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_pdf(title, headers, rows, filename=None, orientation='portrait'):
    buffer = io.BytesIO()
    pagesize = landscape(A4) if orientation == 'landscape' or len(headers) > 8 else A4

    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=6,
        textColor=colors.HexColor('#164B60')
    )
    elements.append(Paragraph(title, title_style))

    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#718096')
    )
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style))
    elements.append(Spacer(1, 8 * mm))

    table_data = [headers]
    for row in rows:
        table_data.append([str(cell) if cell is not None else '' for cell in row])

    page_width = pagesize[0] - 30 * mm
    num_cols = len(headers)
    col_width = page_width / num_cols if num_cols > 0 else page_width

    table = Table(table_data, repeatRows=1)

    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B6B93')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F7FAFC')))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel(title, headers, rows, filename=None):
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else 'Report'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B6B93', end_color='1B6B93', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='164B60')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) if headers else 1)

    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    date_cell.font = Font(name='Calibri', size=9, color='718096')

    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    data_alignment = Alignment(vertical='center', wrap_text=True)

    for row_idx, row_data in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_alignment
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

    wb.save(buffer)
    buffer.seek(0)
    return buffer
